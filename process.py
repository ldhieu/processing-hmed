#!/usr/bin/env python

import openpyxl
import datetime
import asyncio
import argparse
import pandas
from databases import Database

async def database_create_and_get():
    # khởi tạo database tạm
    database = Database("sqlite:///tmp.db")
    await database.connect()
    # tạo table mới, trống dữ liệu
    await database.execute("PRAGMA strict=ON;")
    await database.execute("DROP TABLE IF EXISTS drug_stock;")
    query = """
    CREATE TABLE IF NOT EXISTS drug_stock
    (
    province_name text not null,
    facility_name text not null,
    drug_name text not null,
    drug_source text not null,
    drug_uom text,
    report_year integer not null check( report_year>=2000 and report_year<=2100 ),
    is_monthly integer check(is_monthly = 0 or is_monthly = 1),
    report_month integer check( (is_monthly = 1 and report_month <= 13 and report_month >= 1)
                                or
                                (is_monthly = 0 and report_month = 0)
                              ),
    report_quarter integer check( (is_monthly = 1 and report_quarter = 0)
                                  or
                                  (is_monthly = 0 and report_quarter >= 1 and report_quarter <= 4)
                                ),
    ton_dau_ky integer not null,
    nhap_dinh_ky integer not null,
    nhap_khac integer not null,
    xuat_cho_benh_nhan_trong_ky integer not null,
    xuat_dieu_chuyen_trong_ky integer not null,
    hu_hao integer not null,
    ton_cuoi_ky integer not null
    );"""
    await database.execute(query)
    return database


def get_list_of_drugname_rows(worksheet):
    """
    Hàm để lấy các ô dữ liệu trong Worksheet có tên Thuốc, và xử lý tên Thuốc
    :param worksheet: là biến đặc biệt trữ Worksheet (của openpyxl)
    :return: list các tuple (Cell, tên_thuốc)
    """
    result = list()
    # chạy toàn bộ dòng của worksheet
    for row in worksheet.iter_rows():
        # biến 'c' là ô đầu tiên (A1, B1, ...) của từng dòng
        c = row[0]
        # biến 'v' là value của ô 'c'
        v = str(c.value)
        # nếu value không có dấu gạch ngang "-" ở đầu thì skip
        if v[0] != "-":
            continue
        # nếu value có dấu gạch ngang "-" ở đầu thì ta có dòng thuốc
        else:
            # bỏ "- " đầu nhưng giữ "-" giữa string
            v_real = "-".join(v.split("-")[1:]).strip()
            # gắn vào array trả về
            result.append((c, v_real))
    return result


def process_excel_file(filename):
    """
    Hàm để xử lý file excel cho ra row để gắn vào Db
    :param filename: là tên file relative với code ví dụ "resources/dongnai-cdc-2021-m01.xlsx"
    :return: là list các row để đưa vào sqlite db (xem trên)
    """
    result = list()
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb["Sheet1"]
    # Tên tỉnh
    k_province_name = ws["A1"].value.split(":")[1].split("/")[1].strip()
    # Tên cơ sở
    k_facility_name = ws["A2"].value.split(":")[1].strip()
    # Từ ngày
    k_date_from = datetime.datetime.strptime(
        ws["A7"].value.split("(")[1].split(")")[0].split("-")[0].strip(), "%d/%m/%Y"
    )
    # Đến ngày
    k_date_to = datetime.datetime.strptime(
        ws["A7"].value.split("(")[1].split(")")[0].split("-")[1].strip(), "%d/%m/%Y"
    )
    # Flag Dữ liệu Tháng (nếu True là Tháng, nếu false là Quý)
    f_monthly = k_date_from.month == k_date_to.month
    if not f_monthly:
        # Quý (1,2,3,4) báo cáo
        k_quarter = ((k_date_from.month - 1) // 3) + 1
    else:
        # Tháng (1,2,...,12) báo cáo
        k_month = k_date_from.month
    # Năm báo cáo (YYYY)
    k_year = k_date_from.year
    # Bắt đầu Logic chạy lấy row dữ liệu
    p_thuoc = get_list_of_drugname_rows(ws)
    prev_x = None
    for x in p_thuoc:
        # Tìm điểm bắt đầu
        if prev_x is None:
            prev_x = x
            continue
        else:
            # Tên thuốc
            ten_thuoc = prev_x[1]
            # Tìm các row từ điểm bắt đầu prev_x tới điểm kết thúc (x)
            for row in ws.iter_rows(min_row=prev_x[0].row, max_row=x[0].row):
                # Bỏ row đầu ...
                if row[0].row == prev_x[0].row:
                    continue
                # ... bỏ row đuôi ...
                elif row[0].row == x[0].row:
                    continue
                # ... và bỏ row trống dữ liệu
                elif row[1].value is None:
                    continue
                # Ta có row có dữ liệu
                else:
                    # Fill dữ liệu để đưa vào sqlite
                    data_row = {
                        "province_name": k_province_name,
                        "facility_name": k_facility_name,
                        "drug_name": ten_thuoc,
                        "drug_source": row[1].value.strip(), # cột B
                        "drug_uom": row[3].value.strip(), # cột D
                        "report_year": k_year,
                        "is_monthly": int(f_monthly),
                        "report_month": None, # lát điền sau theo logic f_monthly
                        "report_quarter": None, # lát điền sau theo logic f_monthly
                        "ton_dau_ky": row[7].value, # cột H
                        "nhap_dinh_ky": row[8].value, # cột I
                        "nhap_khac": row[9].value, # cột J
                        "xuat_cho_benh_nhan_trong_ky": row[10].value, # cột K
                        "xuat_dieu_chuyen_trong_ky": row[11].value, # cột L
                        "hu_hao": row[12].value, # cột M
                        "ton_cuoi_ky": row[13].value # cột N
                        }
                    # nếu flag dữ liệu tháng
                    if f_monthly:
                        # điền tháng, để quý None (NULL)
                        data_row["report_month"] = k_month
                    else:
                        # điền quý, để tháng None (NULL)
                        data_row["report_quarter"] = k_quarter
                # thêm vào cục dữ liệu trả về
                result.append(data_row)
            prev_x = x
    # trả dữ liệu là array of rows
    return result

if __name__ == "__main__":
    db = asyncio.run(database_create_and_get())
    parser = argparse.ArgumentParser(description="Process HMED exported XLSX files (convert XLS to XLSX before running)")
    parser.add_argument("files", metavar="file", type=str, nargs="+",
                        help="HMED excel files to process")
    args = parser.parse_args()
    for file in args.files:
        vals = process_excel_file(file)
        query = """INSERT INTO drug_stock(
        province_name, facility_name, drug_name, drug_source, drug_uom, report_year, is_monthly, report_month, report_quarter, ton_dau_ky, nhap_dinh_ky, nhap_khac, xuat_cho_benh_nhan_trong_ky, xuat_dieu_chuyen_trong_ky, hu_hao, ton_cuoi_ky
        ) VALUES (
        :province_name, :facility_name, :drug_name, :drug_source, :drug_uom, :report_year, :is_monthly, :report_month, :report_quarter, :ton_dau_ky, :nhap_dinh_ky, :nhap_khac, :xuat_cho_benh_nhan_trong_ky, :xuat_dieu_chuyen_trong_ky, :hu_hao, :ton_cuoi_ky);"""
        asyncio.run(db.execute_many(query=query, values=vals))
    m_query = """
        SELECT province_name, facility_name, drug_name, drug_source, drug_uom, report_year,
        report_month,
        SUM(ton_dau_ky),
        SUM(nhap_dinh_ky),
        SUM(nhap_khac),
        SUM(xuat_cho_benh_nhan_trong_ky),
        SUM(xuat_dieu_chuyen_trong_ky),
        SUM(hu_hao),
        SUM(ton_cuoi_ky)
        FROM drug_stock
        WHERE is_monthly = 1
        GROUP BY province_name, facility_name, report_year, report_month,
        drug_name, drug_source, drug_uom
        ORDER BY 1, 2, 6, 7, 4, 3, 5 ASC;
        """
    q_query = """
        SELECT province_name, facility_name, drug_name, drug_source, drug_uom, report_year,
        report_quarter,
        SUM(ton_dau_ky),
        SUM(nhap_dinh_ky),
        SUM(nhap_khac),
        SUM(xuat_cho_benh_nhan_trong_ky),
        SUM(xuat_dieu_chuyen_trong_ky),
        SUM(hu_hao),
        SUM(ton_cuoi_ky)
        FROM drug_stock
        WHERE is_monthly = 0
        GROUP BY province_name, facility_name, report_year, report_quarter,
        drug_name, drug_source, drug_uom
        ORDER BY 1, 2, 6, 7, 4, 3, 5 ASC;
        """
    m_rows = asyncio.run(db.fetch_all(query=m_query))
    data_m = pandas.DataFrame(data=m_rows, columns=[
        "Ten_tinh", "Ten_co_so", "Ten_thuoc", "Nguon_thuoc", "Don_vi_tinh",
        "Nam", "Thang",
        "Ton_dau_ky", "Nhap_dinh_ky", "Nhap_khac", "Xuat_cho_benh_nhan_trong_ky",
        "Xuat_dieu_chuyen_trong_ky", "Hu_hao", "Ton_cuoi_ky"
        ])
    data_m.to_excel("output/monthly_all_data.xlsx", index=False)
    q_rows = asyncio.run(db.fetch_all(query=q_query))
    data_q = pandas.DataFrame(data=q_rows, columns=[
        "Ten_tinh", "Ten_co_so", "Ten_thuoc", "Nguon_thuoc", "Don_vi_tinh",
        "Nam", "Quy",
        "Ton_dau_ky", "Nhap_dinh_ky", "Nhap_khac", "Xuat_cho_benh_nhan_trong_ky",
        "Xuat_dieu_chuyen_trong_ky", "Hu_hao", "Ton_cuoi_ky"
        ])
    data_q.to_excel("output/quarterly_all_data.xlsx", index=False)
